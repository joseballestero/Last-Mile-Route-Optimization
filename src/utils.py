"""This module contains utility functions for the VRPTWDO project."""

import copy
import random
import time
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from matplotlib.backends.backend_pdf import PdfPages
from models import *
from sklearn.cluster import KMeans
import matplotlib.dates as mdates
from datetime import datetime, timedelta
import matplotlib.cm as cm

def read_problem_file(file_path: str):
    """
    Reads a problem file and returns a Problem instance.

    Parameters:
        file_path : str
            File path.

    Returns:
        problem : Problem
            Problem instance.
    """

    with open(file_path, 'r') as f:
        lines = f.readlines()

    num_vehicles = int(lines[1].split()[0])
    truck_cost = int(lines[1].split()[1])
    truck_speed = int(lines[1].split()[2])
    km_cost = int(lines[1].split()[3])
    if '.' in lines[1].split()[4] and '.' in lines[1].split()[5]:
        depot_location = Location(float(lines[1].split()[4]), float(lines[1].split()[5]))
    else:
        depot_location = Location(int(lines[1].split()[4]), int(lines[1].split()[5]))
    depot_tw = tuple(map(int, lines[1].split()[6:8]))

    fleet = []
    for i in range(num_vehicles):
        new_truck = Truck(truck_cost)
        fleet.append(new_truck)
    depot = Depot(depot_location, fleet, depot_tw)

    customers = []
    delivery_points = []
    orders = []
    current_order_id = 'O1'
    current_order = None

    for line in lines[3:]:
        strings = line.split()
        strings =  [string.strip() for string in strings]

        if line[0] == 'L' and not 'LOCKER_ID' in line:
            if '.' in strings[1] and '.' in strings[2]:
                loc = Location(float(strings[1]), float(strings[2]))
            else:
                loc = Location(int(strings[1]), int(strings[2]))
            time_to_depot = loc.distance(depot_location) / truck_speed * 60
            new_locker = DeliveryPoint(DeliveryType.LOCKER, loc, time_to_depot, [], int(strings[3]), strings[0])
            delivery_points.append(new_locker)

        elif line[0] == 'S' and not 'SHOP_ID' in line:
            tws = []
            tw = list(map(int, strings[3:5]))
            tw.append(1.0)
            tws.append(tuple(tw))
            if strings[5] != '-':
                tw = list(map(int, strings[5:7]))
                tw.append(1.0)
                tws.append(tuple(tw))
            if '.' in strings[1] and '.' in strings[2]:
                loc = Location(float(strings[1]), float(strings[2]))
            else:
                loc = Location(int(strings[1]), int(strings[2]))
            time_to_depot = loc.distance(depot_location) / truck_speed * 60
            new_shop = DeliveryPoint(DeliveryType.SHOP, loc, time_to_depot, tws, int(strings[7]), strings[0])
            delivery_points.append(new_shop)

        elif line[0] == 'O' and not 'ORDER_ID' in line:
            new_order_id = strings[0]
            new_delivery_option = None

            if strings[4][0] == 'L' or strings[4][0] == 'S':
                dp = next((dp for dp in delivery_points if dp.id == strings[4]), None)
                new_delivery_option = (dp, int(strings[3]))
            elif strings[4][0] == 'H':
                tws = []
                str_tws_probs = strings[7:]
                for str in str_tws_probs:
                    str = str.replace(';', '')
                    nums = str.split(',')
                    tws.append((int(nums[0]), int(nums[1]), float(nums[2])))
                if '.' in strings[5] and '.' in strings[6]:
                    loc = Location(float(strings[5]), float(strings[6]))
                else:
                    loc = Location(int(strings[5]), int(strings[6]))
                time_to_depot = loc.distance(depot_location) / truck_speed * 60
                new_home = PersonalPoint(loc, time_to_depot, False, tws, [], id=strings[4])
                delivery_points.append(new_home)
                new_delivery_option = (new_home, int(strings[3]))

            if new_order_id == current_order_id:
                if current_order is None:
                    new_customer = Customer('C' + current_order_id[1:])
                    customers.append(new_customer)
                    current_order = Order(new_customer, int(strings[2]), int(strings[1]), -1, [], 'O' + current_order_id[1:])
                current_order.add_delivery_option(new_delivery_option)
            else:
                orders.append(current_order)
                current_order_id = new_order_id
                new_customer = Customer('C' + current_order_id[1:])
                customers.append(new_customer)
                current_order = Order(new_customer, int(strings[2]), int(strings[1]), -1, [], 'O' + current_order_id[1:])
                current_order.add_delivery_option(new_delivery_option)

    orders.append(current_order)

    problem = Problem(depot, truck_speed, truck_cost, km_cost, customers, delivery_points, orders)

    return problem


def plot_problem(problem: Problem, file_name: str = 'problem', loc_names: bool = True):
    """
    Plots a problem.

    Parameters:
        problem : Problem
            Problem instance.
        file_name : str
            File name.
    """

    lockers = [dp for dp in problem.delivery_points if dp.delivery_type == DeliveryType.LOCKER]
    shops = [dp for dp in problem.delivery_points if dp.delivery_type == DeliveryType.SHOP]
    home_delivery_points = [dp for dp in problem.delivery_points if dp.delivery_type == DeliveryType.HOME]

    plt.clf()
    plt.rcParams['figure.figsize'] = plt.rcParamsDefault['figure.figsize']

    marker_size = 50
    if len(problem.delivery_points) > 500:
        marker_size = 10
        plt.figure(figsize=(8, 6))

    plt.scatter([dp.loc.x for dp in lockers], [dp.loc.y for dp in lockers], color='#7C92F3', marker='p', s=marker_size, label='Lockers')
    plt.scatter([dp.loc.x for dp in shops], [dp.loc.y for dp in shops], color='#86E9AC', marker='H', s=marker_size, label='Shops')
    plt.scatter([dp.loc.x for dp in home_delivery_points], [dp.loc.y for dp in home_delivery_points], color='#F6A2A8', marker='^', s=marker_size, label='Home Delivery Points')
    plt.scatter(problem.depot.loc.x, problem.depot.loc.y, color='black', marker='D', s=2*marker_size, label='Depot')

    if loc_names:
        name_sep = 0.5
        if plt.xlim()[1] - plt.xlim()[0] < 30: name_sep = 0.15

        for dp in problem.delivery_points:
            plt.text(dp.loc.x + name_sep, dp.loc.y, str(dp.id), fontsize=6, ha='left', va='bottom')

    plt.xlabel('X Coordinate')
    plt.ylabel('Y Coordinate')
    plt.title('Problem Locations')
    plt.grid(True)
    plt.savefig(file_name + '_plot.pdf', format='pdf')


def get_random_solution(problem: Problem):
    """
    Returns a random solution for the tabu search algorithm.

    Parameters:
        problem : Problem
            Problem instance.

    Returns:
        solution : list
            Vector solution. List of (Order_ID, DeliveryPoint_ID) tuples.
    """
    random_solution = problem.create_list_of_options()
    random.shuffle(random_solution)

    return random_solution


def fitness(problem: Problem, cost_weight: float, priority_weight: float, cost: float, priority: list, not_served_count: int = 0):
    """
    Returns the fitness of a solution for optimization algorithms.

    Parameters:
        problem : Problem
            Problem instance.
        cost_weight : float
            Distance weight.
        priority_weight : float
            Priority weight.
        miss_prob_weight: float
            Miss probability weight.
        cost : float
            Total cost of the solution.
        priority : float
            Total priority of the solution.
        miss_prob: float
            Mean probability of missed deliveries.
        not_served_count : int
            Number of orders not served.
    Returns:
        fitness : float
            Fitness of the solution.
    """

    max_priority = get_max_priority(problem) 
    N = len(priority) #La min priority es el número de ordenes (es decir, todas las entregas de prioridad 1)
    #normalized_priority = 1 - (priority / max_priority)
    normalized_priority = np.sum(3 - np.array(priority)) / (2 * N)
    
    # Imprimir las variables antes del cálculo
    # print("=== Variables utilizadas para el cálculo de fitness ===")
    # print(f"Coste total: {cost}")
    # print(f"Prioridad total: {priority}")
    # print(f"Órdenes no servidas: {not_served_count}")
    # print(f"Coste normalizado: {normalized_cost}")
    # print(f"Prioridad normalizada: {normalized_priority}")
    # print(f"Peso del coste (alpha): {cost_weight}")
    # print(f"Peso de la prioridad (beta): {priority_weight}")
    # print(f"Órdenes no servidas: {not_served_count}")   
    fitness_value = (cost_weight + priority_weight * normalized_priority + not_served_count * 0.01) * 100

    return normalized_priority


def save_solution(solution: list, file_name: str, stats: list = []):
    """
    Saves a solution to a file.

    Parameters:
        solution : list
            Vector solution. List of (Order_ID, DeliveryPoint_ID) tuples.
        file_name : str
            File name.
    """
    min_values = [s["min"] for s in stats]
    np.savez(file_name, solution=np.array(solution), stats=np.array(min_values))


def read_solution_file(file_path: str, problem: Problem):
    """
    Reads a solution file and returns a solution as a vector.

    Parameters:
        file_path : str
            File path.
        problem : Problem
            Problem instance.

    Returns:
        solution : list
            Vector solution. List of (Order_ID, DeliveryPoint_ID) tuples.
    """

    solution, stats = [], []
    data = np.load(file_path)
    if file_path.endswith(".npz"):
        if len(data["stats"]):
            solution = [tuple(row) for row in data["solution"]]
            stats = list(data["stats"])
        else:
            solution = [tuple(row) for row in data["solution"]]
    elif file_path.endswith(".npy"):
        solution = [tuple(row) for row in data]
    return solution, stats




def create_solution(problem: Problem, solution: list):
    """
    Returns a solution object.

    Parameters:
        problem : Problem
            Problem instance.
        solution : list
            Vector solution. List of delivery options.

    Returns:
        solution : Solution
            Solution instance.
    """
    start = time.time()

    routes = []
    served = {order.id: 0 for order in problem.orders}
    capacity = copy.deepcopy(problem.dict_capacity)

    truck = Truck(problem.truck_cost)
    route = Route(truck)
    route_as_list = [(None,'DEPOT')]
    route.start_time = [problem.dict_twe['DEPOT']]
    route_arrival = [problem.dict_twe['DEPOT']]
    route_initial_time = [problem.dict_twe['DEPOT']]
    route_departure = [problem.dict_twe['DEPOT']]
    distance = 0
    total_time = 0
    priority = 0

    ind = 0
    while ind < len(solution):
        option = solution[ind]
        order, dp = option
        last_dp = route_as_list[-1][1]

        if not served[order] and capacity[dp] != 0:
            arrival_time = route_departure[-1] + problem.dict_distance[(last_dp, dp)] * 60 / problem.truck_speed
            dp_opts = []
            if isinstance(problem.dict_twe[dp], list) and isinstance(problem.dict_twl[dp], list):
                for te, tl in zip(problem.dict_twe[dp], problem.dict_twl[dp]):
                    initial_time = max(arrival_time, te)
                    twe, twl = te, tl
                    if initial_time <= tl:
                        dp_opts.append((initial_time, twe, twl))
            else:
                twe, twl = problem.dict_twe[dp], problem.dict_twl[dp]
                initial_time = max(arrival_time,  twe)
                dp_opts.append((initial_time, twe, twl))

            delivery_time = 0
            if dp != last_dp:
                delivery_time += problem.dict_delivery_time['DEFAULT']
            if 'H' == dp[0]:
                delivery_time += problem.dict_delivery_time['HOME']
            elif 'L' == dp[0]:
                delivery_time += problem.dict_delivery_time['LOCKER']
            elif 'S' == dp[0]:
                delivery_time += problem.dict_delivery_time['SHOP']

            if len(dp_opts) > 0:
                selected = False
                for initial_time, twe, twl in dp_opts:
                    departure_time = initial_time + delivery_time
                    arrival_time_depot = departure_time + problem.dict_distance[(dp, 'DEPOT')] * 60 / problem.truck_speed
                    if arrival_time <= twl and arrival_time_depot <= problem.dict_twl['DEPOT']:
                        distance += problem.dict_distance[(last_dp, dp)]
                        total_time += problem.dict_distance[(last_dp, dp)] * 60 / problem.truck_speed
                        priority += problem.dict_priority[option]

                        dp = problem.find_dp_by_id(dp)
                        order_obj = problem.find_order_by_id(order)
                        stop = Stop(dp, [order_obj], initial_time)
                        route.stops.append(stop)

                        route_as_list.append(option)
                        route_arrival.append(arrival_time)
                        route_initial_time.append(initial_time)
                        route_departure.append(departure_time)
                        served[order] = 1
                        capacity[dp.id] -= 1
                        ind += 1
                        selected = True
                        break
                if not selected:
                    if len(route_as_list) > 1 and last_dp != 'DEPOT':
                        distance += problem.dict_distance[(last_dp, 'DEPOT')]
                        total_time += problem.dict_distance[(last_dp, 'DEPOT')] * 60 / problem.truck_speed
                        route_copy = copy.deepcopy(route)
                        routes.append(route_copy)
                        new_truck = Truck(problem.truck_cost)
                        route.clear()
                        route.truck = new_truck
                        route_as_list = copy.deepcopy([(None,'DEPOT')])
                        route.start_time = [problem.dict_twe['DEPOT']]
                        route_arrival = [problem.dict_twe['DEPOT']]
                        route_initial_time = [problem.dict_twe['DEPOT']]
                        route_departure = [problem.dict_twe['DEPOT']]
                    else:
                        # DP is not reachable
                        ind += 1
            else:
                ind += 1
        else:
            ind += 1

        if ind + 1 == len(solution):
            distance += problem.dict_distance[(last_dp, 'DEPOT')]
            total_time += problem.dict_distance[(last_dp, 'DEPOT')] * 60 / problem.truck_speed
            route_copy = copy.deepcopy(route)
            routes.append(route_copy)

    not_served_count = sum(1 for x in served.values() if x == 0)
    # print("Not served: ", not_served_count)

    # print("Time (seconds): ", time.time() - start)

    solution = Solution(routes)

    return solution


def evaluate_routes(problem, final_routes):
    """
    Evaluates each route separately and aggregates total results.

    Parameters:
        problem : Problem
            The problem instance containing all relevant data.
        final_routes : list
            List of routes, where each route is a list of (order_id, delivery_point).

    Returns:
        total_priority : float
            Total priority score of completed deliveries.
        total_distance : float
            Total distance covered by all routes.
        total_time : float
            Total time spent on all deliveries.
        not_served_count : int
            Number of orders not served.
        delivery_times : list
            List of delivery times for each route.
    """
    total_priority = []
    total_distance = 0
    total_time = 0
    served_orders = set()
    delivery_times = []
    
    delivery_time_mapping = {
        'H': problem.dict_delivery_time['HOME'],
        'L': problem.dict_delivery_time['LOCKER'],
        'S': problem.dict_delivery_time['SHOP']
    }

    for vehicle_id, route in enumerate(final_routes):
        if not route:
            continue  # Skip empty routes
        
        # Variables específicas de la ruta
        route_priority = 0
        route_distance = 0
        route_time = 0
        route_times = []
        last_dp = 'DEPOT'
        current_time = problem.dict_twe['DEPOT']

        print(f"\nEvaluando Ruta del Vehículo {vehicle_id}")

        for order_id, dp in route:
            if order_id in served_orders:
                continue  # Evita contar un pedido más de una vez

            # Calcular tiempo de viaje y actualizar tiempo actual
            travel_time = problem.dict_distance[(last_dp, dp)] * 60 / problem.truck_speed
            arrival_time = current_time + travel_time

            # Verificar si se cumple la ventana de tiempo
            waiting_time = 0
            if isinstance(problem.dict_twe[dp], list):  # Múltiples ventanas
                valid_window_found = False
                for twe, twl in zip(problem.dict_twe[dp], problem.dict_twl[dp]):
                    if twe <= arrival_time <= twl:
                        waiting_time = max(0, twe - arrival_time)
                        valid_window_found = True
                        break
                if not valid_window_found:
                    print(f"🚨 Orden {order_id} en {dp} no se puede servir por restricción de ventana de tiempo.")
                    continue
            else:  # Única ventana de tiempo
                twe, twl = problem.dict_twe[dp], problem.dict_twl[dp]
                waiting_time = max(0, twe - arrival_time)
                if arrival_time > twl:
                    print(f"🚨 Orden {order_id} en {dp} no se puede servir por restricción de ventana de tiempo.")
                    continue

            # Obtener tiempo de entrega y validar que cabe en la ventana
            delivery_time = delivery_time_mapping.get(dp[0], problem.dict_delivery_time['DEFAULT'])
            if arrival_time + waiting_time + delivery_time > twl:
                print(f"🚨 Orden {order_id} en {dp} no se puede servir: entrega fuera de la ventana.")
                continue

            # Actualizar métricas de la ruta
            current_time += waiting_time + delivery_time
            route_distance += problem.dict_distance[(last_dp, dp)]
            route_time += travel_time + waiting_time + delivery_time
            #route_priority += problem.dict_priority[(order_id, dp)]
            total_priority.append(problem.dict_priority[(order_id, dp)])
            
            # Guardar el tiempo de entrega para esta entrega
            route_times.append(round(current_time, 2))

            # Marcar el pedido como entregado
            served_orders.add(order_id)
            last_dp = dp

        # Volver al depósito si la ruta tiene entregas
        if route:
            travel_time_back = problem.dict_distance[(last_dp, 'DEPOT')] * 60 / problem.truck_speed
            current_time += travel_time_back
            route_distance += problem.dict_distance[(last_dp, 'DEPOT')]
            route_time += travel_time_back

        # Acumular resultados en totales
        #total_priority.append(route_priority)
        
        total_distance += route_distance
        total_time += route_time
        delivery_times.append(route_times)

    # Calcular los pedidos no servidos
    total_orders = len(problem.orders)
    not_served_count = max(0, total_orders - len(served_orders))

    return total_priority, total_distance, total_time, not_served_count, delivery_times




def get_solution_charts(problem: Problem, solution: Solution, file_name: str, stats: list = [], loc_names: bool = True):
    """
    Generates the charts of a solution.

    Parameters:
        problem : Problem
            Problem instance.
        solution : Solution
            Solution instance.
        file_name : str
            File name.
        stats : list
            List of stats.
    """

    pdf = PdfPages(file_name + '.pdf')

    tws_chart = plot_solution_tws(problem, solution)
    pdf.savefig(tws_chart)

    routes_chart = plot_solution_routes(problem, solution, loc_names=loc_names)
    pdf.savefig(routes_chart)

    if len(stats):
        fitness = plot_fitness_evolution(stats)
        pdf.savefig(fitness)

    pdf.close()


def plot_solution_tws(problem: Problem, solution: Solution):

    # generate a gantt bar chart with the time windows of each stop, being the x axis the time and the y axis the stops
    plt.clf()
    plt.rcParams['figure.figsize'] = plt.rcParamsDefault['figure.figsize']
    y = 0
    colors = ["#ff0e41","#d500e0","#ff8b1f","#01befe","#7be000","#500aff","#e0c200","#20ac56", "#8b8c89", "#005fac"]
    light_colors = ["#ff9fb4","#f98dff","#ffd1a5","#9ae5fe","#ccff8d","#b99dff","#fff08d","#98ecb8", "#d0d1d0", "#78c2ff"]
    routes = copy.deepcopy(solution.routes)
    routes.reverse()
    for route in routes:
        r_index = solution.routes.index(route)
        color_index = r_index % len(colors)
        stop_exec_times = []
        stops = copy.deepcopy(route.stops)
        stops.reverse()
        for i in range(len(stops)):
            tws = stops[i].dp.time_windows
            if stops[i].dp != stops[i - 1].dp:
                y += 1
                if len(tws) == 0:
                    width = problem.depot.time_window[1] - problem.depot.time_window[0]
                    plt.barh(y, width, left=problem.depot.time_window[0], height=0.5, color=light_colors[color_index], alpha=0.5)
                else:
                    for tw in tws:
                        plt.barh(y, tw[1] - tw[0], left=tw[0], height=0.5, color=light_colors[color_index], alpha=0.5)

            plt.scatter(stops[i].exec_time, y, color=colors[color_index], marker=".", s=30)
            stop_exec_times.append((stops[i].exec_time, y))

        for i in range(len(stop_exec_times) - 1):
            plt.plot([stop_exec_times[i][0], stop_exec_times[i + 1][0]], [stop_exec_times[i][1], stop_exec_times[i + 1][1]], color=colors[color_index], linestyle='dashed', linewidth=1)


    plt.xlabel('Time')
    plt.ylabel('Delivery Points')
    plt.yticks([])  #  remove y axis ticks
    plt.title('Time Windows Chart')
    fig = plt.gcf()

    return fig



def plot_solution_routes(problem: Problem, solution: Solution, loc_names: bool = True):
    """
    Plots a solution.

    Parameters:
        problem : Problem
            Problem instance.
        solution : Solution
            Solution instance.
        file_name : str
            File name.
    """

    lockers = [dp for dp in problem.delivery_points if dp.delivery_type == DeliveryType.LOCKER]
    shops = [dp for dp in problem.delivery_points if dp.delivery_type == DeliveryType.SHOP]
    home_delivery_points = [dp for dp in problem.delivery_points if dp.delivery_type == DeliveryType.HOME]

    plt.clf()
    plt.rcParams['figure.figsize'] = plt.rcParamsDefault['figure.figsize']

    marker_size = 50
    if len(problem.delivery_points) > 500:
        marker_size = 10
        plt.figure(figsize=(8, 6))

    plt.scatter([dp.loc.x for dp in lockers], [dp.loc.y for dp in lockers], color='#7C92F3', marker='p', s=marker_size, label='Lockers')
    plt.scatter([dp.loc.x for dp in shops], [dp.loc.y for dp in shops], color='#86E9AC', marker='H', s=marker_size, label='Shops')
    plt.scatter([dp.loc.x for dp in home_delivery_points], [dp.loc.y for dp in home_delivery_points], color='#F6A2A8', marker='^', s=marker_size, label='Home Delivery Points')
    plt.scatter(problem.depot.loc.x, problem.depot.loc.y, color='black', marker='D', s=2*marker_size, label='Depot')

    colors = ["#ff0e41","#d500e0","#ff8b1f","#01befe","#7be000","#500aff","#e0c200","#20ac56", "#8b8c89", "#005fac"]

    for route in solution.routes:
        r_index = solution.routes.index(route)
        color_index = r_index % len(colors)
        for i in range(len(route.stops)):
            if i == 0:
                plt.plot([problem.depot.loc.x, route.stops[i].dp.loc.x], [problem.depot.loc.y, route.stops[i].dp.loc.y], color=colors[color_index], linestyle='dashed', linewidth=1)
            elif i == len(route.stops) - 1:
                plt.plot([route.stops[i - 1].dp.loc.x, route.stops[i].dp.loc.x], [route.stops[i - 1].dp.loc.y, route.stops[i].dp.loc.y], color=colors[color_index], linestyle='dashed', linewidth=1)
                plt.plot([route.stops[i].dp.loc.x, problem.depot.loc.x], [route.stops[i].dp.loc.y, problem.depot.loc.y], color=colors[color_index], linestyle='dashed', linewidth=1)
            else:
                plt.plot([route.stops[i - 1].dp.loc.x, route.stops[i].dp.loc.x], [route.stops[i - 1].dp.loc.y, route.stops[i].dp.loc.y], color=colors[color_index], linestyle='dashed', linewidth=1)

    if loc_names:
        for dp in problem.delivery_points:
            plt.text(dp.loc.x * 1.005, dp.loc.y, str(dp.id), fontsize=6, ha='left', va='bottom')

    plt.xlabel('X Coordinate')
    plt.ylabel('Y Coordinate')
    plt.title('Locations and Routes')
    fig = plt.gcf()

    return fig


def plot_fitness_evolution(stats_list: list):
    if isinstance(stats_list[0], float):
        stats_list = [{"min": s} for s in stats_list]

    # max_list = [s["max"] for s in stats_list]
    min_list = [s["min"] for s in stats_list]
    # avg_list = [s["avg"] for s in stats_list]

    plt.clf()
    # plt.figure(figsize=(10, 5))
    plt.plot(min_list, color="#ff0e41", label='Min fitness')
    # plt.plot(max_list, color="#005fac", label='Max')
    # plt.plot(avg_list, color="#e0c200", label='Avg')

    plt.xlabel('Iterations')
    plt.ylabel('Fitness value')
    plt.title('Fitness evolution')
    plt.legend()
    fig = plt.gcf()

    return fig


def get_max_cost(problem: Problem):
    """
    Returns the estimated maximum cost of the problem.

    Parameters:
        problem : Problem
            Problem instance.

    Returns:
        max_cost : float
            Maximum cost of the problem.
    """

    max_cost = 0.0
    initial_options = problem.create_list_of_options()
    priority, distance, routes, total_time, not_served_count, delivery_times = eval_solution(problem, initial_options)

    cost = distance * problem.km_cost + len(routes) * problem.truck_cost
    max_cost = 1.3 * cost
    # print("Max cost: ", max_cost)

    return max_cost



def export_to_excel(wb, name, *args):
    
    # Seleccionar la hoja activa (por defecto, la primera hoja)
    ws = wb.active
    
    # Obtener la fila actual para escribir los datos
    current_row = ws.max_row + 1
    
    # Iterar sobre los argumentos pasados y escribir cada uno en una columna diferente
    for i, arg in enumerate(args, start=1):
        # Obtener la letra de la columna correspondiente
        col_letter = get_column_letter(i)
        # Escribir el valor del argumento en la celda adecuada
        ws[col_letter + str(current_row)] = arg
    
    # Guardar el libro de trabajo en el archivo especificado
    wb.save(name)


def __get_option_id(option: tuple):
    """
    Returns the id of a delivery option.

    Parameters:
        option : tuple
            ((DeliveryPoint | PersonalPoint, priority), Order) tuple.

    Returns:
        option_id : str
            Id of the delivery option.
    """

    dp_id = id(option[0][0].id)
    order_id = id(option[1].id)
    return f"{dp_id}-{order_id}"

    
    
def plot_clusters(destinos, labels):
    
    #Representar clusters
    # Convertir las ubicaciones a formato numpy para facilidad de uso con matplotlib
    coords = np.array([[point[0], point[1]] for point in destinos])
    print(f'coords shape: {coords.shape}')
    
    # Representar los puntos de entrega con diferentes colores según los clusters
    plt.figure(figsize=(10, 8))
    unique_labels = np.unique(labels)
    
    # Utilizar un mapa de colores con suficiente diversidad para los clusters
    colors = plt.cm.get_cmap('tab20', len(unique_labels))
    
    for label in unique_labels:
        cluster_points = coords[labels == label]
        print(f'Cluster {label} points shape: {cluster_points.shape}')
        
        plt.scatter(cluster_points[:, 0], cluster_points[:, 1], label=f'Cluster {label}', color=colors(label / len(unique_labels)))
    
        
    plt.title('Delivery Points Clusters')
    plt.xlabel('X Coordinate')
    plt.ylabel('Y Coordinate')
    plt.legend()
    plt.show()



#Convert locations XY from delivery_points to a tupla.
def convert_locations(problem):
    locations = []
    for order in problem.orders:
        for option in order.delivery_options:
            dp, priority = option[0], option[1]
            locations.append([dp.loc.x, dp.loc.y])
    return np.array(locations)          
    

# Elbow method to determine optimal k
def optimal_k(locations, max_k=10):
    
    inertia = []
    for k in range(1, max_k + 1):
        kmeans = KMeans(n_clusters=k, random_state=42).fit(locations)
        inertia.append(kmeans.inertia_)
    
    plt.plot(range(1, max_k + 1), inertia, 'bx-')
    plt.xlabel('Number of clusters (k)')
    plt.ylabel('Inertia')
    plt.title('Elbow Method for Optimal k')
    plt.show()

    # Elige k basándote en el gráfico
    return int(input("Enter the optimal number of clusters (k): "))


def plot_clusters(locations, labels):
    plt.scatter(locations[:, 0], locations[:, 1], c=labels, s=50, cmap='viridis')
    plt.xlabel('X Coordinate')
    plt.ylabel('Y Coordinate')
    plt.title('Clusters of Delivery Points')
    plt.show()




def two_opt(route, problem):
    def calculate_distance(route):
        total_distance = 0
        for i in range(len(route) - 1):
            total_distance += problem.dict_distance[(route[i][1], route[i+1][1])]
        return total_distance

    best_route = route
    best_distance = calculate_distance(route)
    
    improved = True
    while improved:
        improved = False
        for i in range(1, len(route) - 2):
            for j in range(i + 1, len(route)):
                if j - i == 1: continue
                new_route = route[:i] + route[i:j][::-1] + route[j:]
                new_distance = calculate_distance(new_route)
                if new_distance < best_distance:
                    best_route = new_route
                    best_distance = new_distance
                    improved = True
        route = best_route
    return best_route


def generate_initial_solution(problem, num_clusters):
    coords = np.array([[point.x, point.y] for point in problem.dict_xy.values()])
    kmeans = KMeans(n_clusters=num_clusters, random_state=0).fit(coords)
    labels = kmeans.labels_

    clusters = {i: [] for i in range(num_clusters)}
    for point, label in zip(problem.list_of_options, labels):
        clusters[label].append(point)

    initial_solution = []
    for cluster_id, points in clusters.items():
        sorted_points = sorted(points, key=lambda x: x[0].twe)
        route = []
        current_capacity = problem.truck_capacity
        for point in sorted_points:
            if current_capacity > 0:
                route.append((point[1].id, point[0].id))
                current_capacity -= 1
            else:
                initial_solution.append(route)
                route = [(point[1].id, point[0].id)]
                current_capacity = problem.truck_capacity - 1
        if route:
            initial_solution.append(route)

    return initial_solution


def get_max_priority(problem: Problem):
    """
    Returns the maximum priority of the problem.

    Parameters:
        problem : Problem
            Problem instance.

    Returns:
        max_priority : float
            Maximum priority of the problem.
    """

    max_priority = 0

    for order in problem.orders:
        order_priority = len(order.delivery_options)
        max_priority += order_priority

    return max_priority



def apply_kmeans(problem, num_vehicles, alpha, beta, gamma, max_time_per_vehicle=480, plot_name="kmeans_clusters"):
    """
    Apply K-Means clustering to assign delivery points to different vehicles based on location.

    Parameters:
        problem : Problem
            The problem instance containing delivery points.
        num_vehicles : int
            Number of vehicles (trucks) available.
        alpha: float
            Weight for distance in the cost function.
        beta: float
            Weight for waiting time in the cost function.
        gamma: float
            Weight for priority in the cost function.
        max_time_per_vehicle: int
            Maximum time each vehicle can work in minutes (e.g., 480 minutes for 8 hours).

    Returns:
        clustered_solution : dict
            Dictionary with vehicle ID as key and list of delivery points as values.
    """
    delivery_coords = []
    delivery_points = []

    # Collect the coordinates and delivery points for clustering
    for order in problem.orders:
        for option in order.delivery_options:
            dp = option[0]  # Delivery point
            delivery_coords.append([dp.loc.x, dp.loc.y])
            delivery_points.append((order.id, dp.id))

    # Convert delivery coordinates to numpy array for K-Means
    delivery_coords = np.array(delivery_coords)

    # Apply K-Means clustering
    kmeans = KMeans(n_clusters=num_vehicles, random_state=42).fit(delivery_coords)
    labels = kmeans.labels_  # Cluster labels

    # Assign deliveries to vehicles based on clustering
    clustered_solution = {i: [] for i in range(num_vehicles)}
    for i, label in enumerate(labels):
        clustered_solution[label].append(delivery_points[i])

    # Plot clusters with depot if plot_name is provided
    plot_clusters_with_depot(problem, clustered_solution, plot_name)

    return clustered_solution


def elbow_method(problem, max_clusters=10, random_state=42):
    """
    Apply the elbow method to determine the optimal number of clusters for K-Means.

    Parameters:
        problem : Problem
            The problem instance containing delivery points.
        max_clusters : int
            Maximum number of clusters to evaluate.
        random_state : int
            Random state for K-Means initialization.

    Returns:
        None (Plots the elbow curve).
    """
    # Extract delivery point coordinates from problem.dict_xy
    delivery_coords = [coords for dp_id, coords in problem.dict_xy.items() if dp_id != 'DEPOT']
    delivery_coords = np.array(delivery_coords)

    # Calculate the Within-Cluster-Sum of Squares (WCSS) for different cluster numbers
    wcss = []
    for n_clusters in range(1, max_clusters + 1):
        kmeans = KMeans(n_clusters=n_clusters, random_state=random_state)
        kmeans.fit(delivery_coords)
        wcss.append(kmeans.inertia_)

    # Plot the elbow curve
    plt.figure(figsize=(10, 6))
    plt.plot(range(1, max_clusters + 1), wcss, marker='o', linestyle='--', color='b')
    plt.title('Elbow Method for Optimal K')
    plt.xlabel('Number of Clusters (K)')
    plt.ylabel('WCSS (Within-Cluster-Sum of Squares)')
    plt.xticks(range(1, max_clusters + 1))
    plt.grid(True)
    plt.show()

def assign_deliveries_to_vehicles(clustered_solution, problem, num_vehicles, max_time_per_vehicle, alpha, beta, gamma):
    """
    Assigns deliveries to vehicles while respecting capacity constraints and time windows.

    Parameters:
        clustered_solution (dict): Dictionary mapping vehicle IDs to lists of delivery points.
        problem (Problem): The problem instance.
        num_vehicles (int): Number of vehicles available.
        max_time_per_vehicle (int): Maximum time per vehicle in minutes.
        alpha (float): Weight for distance in the cost function.
        beta (float): Weight for waiting time in the cost function.
        gamma (float): Weight for priority in the cost function.

    Returns:
        final_routes (list): List of routes, each containing tuples of (order_id, delivery_point).
        total_waiting_time (float): Total waiting time in minutes.
        current_storage (dict): Dictionary tracking the current storage usage of Lockers and Shops.
    """
    MAX_WAITING_TIME = 20  # Maximum waiting time allowed in minutes

    final_routes = []
    served_orders = set()
    total_waiting_time = 0

    # Track current storage usage at each Locker and Shop
    current_storage = {dp_id: 0 for dp_id in problem.dict_capacity.keys()}

    delivery_time_mapping = {
        'H': problem.dict_delivery_time['HOME'],
        'L': problem.dict_delivery_time['LOCKER'],
        'S': problem.dict_delivery_time['SHOP']
    }

    for vehicle_id in range(num_vehicles):
        vehicle_route = []
        current_location = 'DEPOT'
        current_time = problem.dict_twe['DEPOT']
        elapsed_time = 0  # Tiempo trabajado
        delivery_points = clustered_solution[vehicle_id]

        print(f"\nVehículo {vehicle_id} comienza su ruta desde {current_location} a las {current_time:.2f}")

        while delivery_points and elapsed_time < max_time_per_vehicle:
            next_option = None
            min_cost = float('inf')

            for dp_order_id, dp_id in delivery_points:
                if dp_order_id in served_orders:
                    continue

                # Check capacity constraint (only for Lockers 'L' and Shops 'S')
                if dp_id.startswith(('L', 'S')) and dp_id in problem.dict_capacity:
                    if current_storage[dp_id] >= problem.dict_capacity[dp_id]:
                        continue  # Skip this delivery, as the location is full

                # Calcular tiempo de viaje y hora de llegada
                distance = problem.dict_distance[(current_location, dp_id)]
                travel_time = distance * 60 / problem.truck_speed
                arrival_time = current_time + travel_time

                # Verificar ventanas de tiempo
                waiting_time = 0
                if dp_id.startswith('S'):  # Tienda con múltiples ventanas
                    valid_window = False
                    for twe, twl in zip(problem.dict_twe[dp_id], problem.dict_twl[dp_id]):
                        if twe <= arrival_time <= twl:
                            valid_window = True
                            waiting_time = max(0, twe - arrival_time)
                            break
                    if not valid_window:
                        continue
                else:  # Un solo intervalo de tiempo (hogar o locker)
                    twe, twl = problem.dict_twe[dp_id], problem.dict_twl[dp_id]
                    waiting_time = max(0, twe - arrival_time)
                    if arrival_time > twl:
                        continue

                # ❌ Rechazar la entrega si el tiempo de espera es mayor a 40 minutos
                if waiting_time > MAX_WAITING_TIME:
                    continue

                # Obtener tiempo de entrega y validar que cabe en la ventana de tiempo
                delivery_time = delivery_time_mapping.get(dp_id[0], problem.dict_delivery_time['DEFAULT'])
                if arrival_time + waiting_time + delivery_time > twl:
                    continue  # No se puede entregar si se pasa del tiempo máximo

                # Calcular coste
                delivery_priority = problem.dict_priority[(dp_order_id, dp_id)]
                cost = (alpha * distance) + (beta * waiting_time) + (gamma * delivery_priority)

                if cost < min_cost:
                    min_cost = cost
                    next_option = (dp_order_id, dp_id, waiting_time)

            if next_option:
                order_id, dp_id, waiting_time = next_option
                served_orders.add(order_id)
                vehicle_route.append((order_id, dp_id))

                # Actualizar el almacenamiento si es Locker o Tienda
                if dp_id.startswith(('L', 'S')):
                    current_storage[dp_id] += 1

                # Calcular tiempos y actualizar estado
                distance_to_next = problem.dict_distance[(current_location, dp_id)]
                travel_time = distance_to_next * 60 / problem.truck_speed
                delivery_time = delivery_time_mapping.get(dp_id[0], problem.dict_delivery_time['DEFAULT'])

                # Actualizar hora actual y tiempo trabajado
                current_time += travel_time + waiting_time + delivery_time
                elapsed_time += travel_time + waiting_time + delivery_time
                total_waiting_time += waiting_time
                current_location = dp_id

                #print(f"Vehículo {vehicle_id} entrega en {dp_id} a las {current_time:.2f}, tiempo trabajado {elapsed_time:.2f} minutos")

                # Validar tiempo restante para regresar al depósito
                return_time = problem.dict_distance[(current_location, 'DEPOT')] * 60 / problem.truck_speed
                if elapsed_time + return_time > max_time_per_vehicle:
                    print(f"Vehículo {vehicle_id} alcanza el tiempo máximo de trabajo ({max_time_per_vehicle} minutos). Regresando al depósito.")
                    break
            else:
                print(f"Vehículo {vehicle_id} no puede realizar más entregas válidas.")
                break

        # Regresar al depósito
        if current_location != 'DEPOT':
            return_time = problem.dict_distance[(current_location, 'DEPOT')] * 60 / problem.truck_speed
            current_time += return_time
            elapsed_time += return_time
            print(f"Vehículo {vehicle_id} regresa al depósito a las {current_time:.2f}, tiempo total trabajado {elapsed_time:.2f} minutos.")
        
        final_routes.append(vehicle_route)

    return final_routes, total_waiting_time, current_storage



# Evaluar las rutas finales generadas por los vehículos
def evaluate_final_solution(final_routes, problem, alpha, beta, gamma):
    """
    Evaluates the final routes generated by the vehicles and calculates key performance metrics.

    Parameters:
        final_routes (list): List of vehicle routes, each containing tuples (order_id, delivery_point).
        problem (Problem): The problem instance containing relevant delivery data.
        alpha (float): Weight for distance in the cost function.
        beta (float): Weight for waiting time in the cost function.
        gamma (float): Weight for priority in the cost function.

    Returns:
        normalized_priority (float): Normalized priority value for fitness evaluation.
        total_cost (float): Total cost computed based on distance and number of vehicles used.
        total_priority (float): Sum of priority values across all deliveries.
        not_served_count (int): Number of orders that were not successfully delivered.
        total_time (float): Total time spent across all routes.
        total_distance (float): Total distance covered by all vehicles.
    """
    total_distance = 0
    total_time = 0
    total_orders = 0
    not_served_count = 0
    delivery_times = []

    # Combinar todas las rutas de los vehículos en una sola lista para evaluar
    clustered_routes = [delivery for route in final_routes for delivery in route]

    # Usamos eval_solution para evaluar todas las rutas juntas
    priority, distance, total_time, not_served_count, delivery_times = evaluate_routes(problem, final_routes)
    #print(priority)
    # Calcular el coste total (por distancia y vehículos)
    total_cost = distance * problem.km_cost + len(final_routes) * problem.truck_cost

    # Calcular el fitness total
    normalized_priority = fitness(problem, alpha, beta, total_cost, priority, not_served_count)

    # print(f"\nEvaluación completa de la solución:")
    # print(f"Prioridad total: {priority}")
    # print(f"Distancia total: {distance} km")
    # print(f"Tiempo total: {total_time} min")
    # print(f"Órdenes no servidas: {not_served_count}")
    # print(f"Coste total: {total_cost}")
    # print(f"Fitness total: {fitness_value}")

    return normalized_priority, total_cost, np.sum(priority), not_served_count, total_time, distance


def create_initial_solution(problem: Problem, alpha: float = 1, beta: float = 1, gamma: float = 1):
    initial_solution = []
    current_location = 'DEPOT'  # The truck starts at the depot
    served_orders = set()  # Track which orders have been served
    remaining_orders = {order.id for order in problem.orders}  # Orders that are not yet served

    current_time = 0  # To track the current time at each delivery point

    while remaining_orders:
        next_option = None
        min_cost = float('inf')

        # Iterate over all possible delivery options to find the one with the lowest cost
        for order in problem.orders:
            if order.id in served_orders:
                continue  # Skip already served orders

            for option in order.delivery_options:
                dp = option[0]  # delivery point (locker, home, shop, etc.)
                if problem.dict_capacity[dp.id] > 0:  # Only consider points with capacity
                    # Calculate arrival time at this point
                    distance = problem.dict_distance[(current_location, dp.id)]
                    arrival_time = current_time + (distance * 60 / problem.truck_speed)

                    # Check if the time window is a list (multiple windows for stores)
                    if isinstance(problem.dict_twe[dp.id], list):
                        valid_window_found = False
                        for twe, twl in zip(problem.dict_twe[dp.id], problem.dict_twl[dp.id]):
                            waiting_time = max(0, twe - arrival_time)
                            if arrival_time <= twl:
                                # Update current time to the start of this valid window
                                current_time = max(arrival_time, twe)
                                valid_window_found = True
                                break
                        if not valid_window_found:
                            continue  # Skip this option if no valid window found
                    else:
                        # Single time window
                        twe, twl = problem.dict_twe[dp.id], problem.dict_twl[dp.id]
                        waiting_time = max(0, twe - arrival_time)
                        if arrival_time > twl:
                            continue  # Skip if out of time window

                    # Get the priority of this delivery
                    delivery_priority = problem.dict_priority[(order.id, dp.id)]

                    # Calculate the total cost for this option
                    cost = (alpha * distance) + (beta * waiting_time) + (gamma * delivery_priority)

                    # Select the option with the lowest cost
                    if cost < min_cost:
                        min_cost = cost
                        next_option = (order.id, dp.id)

        if not next_option:
            # No valid options found, return to depot
            current_location = 'DEPOT'
            current_time += problem.dict_distance[(current_location, 'DEPOT')] * 60 / problem.truck_speed
            break

        # Update solution with the selected option (order, delivery point)
        order_id, dp_id = next_option
        initial_solution.append(next_option)

        # Mark the order as served
        served_orders.add(order_id)
        remaining_orders.remove(order_id)

        # Update current location and time
        current_location = dp_id

        # If the next delivery point is a locker or shop, group all possible orders
        if dp_id.startswith('L') or dp_id.startswith('S'):  # 'L' for lockers, 'S' for stores
            grouped_orders = [o for o in problem.orders if o.id in remaining_orders and dp_id in [opt[0].id for opt in o.delivery_options]]

            for grouped_order in grouped_orders:
                if problem.dict_capacity[dp_id] > 0:
                    if dp_id.startswith('S'):  # Only check time window for stores
                        arrival_time = current_time + (problem.dict_distance[(current_location, dp_id)] * 60 / problem.truck_speed)
                        if isinstance(problem.dict_twe[dp_id], list):
                            valid_window_found = False
                            for twe, twl in zip(problem.dict_twe[dp_id], problem.dict_twl[dp_id]):
                                if arrival_time <= twl:
                                    current_time = max(arrival_time, twe)  # Update time for the window
                                    valid_window_found = True
                                    break
                            if not valid_window_found:
                                continue
                        else:
                            if arrival_time > problem.dict_twl[dp_id]:
                                continue  # Skip this order if it doesn't fit the time window

                    # Add grouped order to the solution
                    initial_solution.append((grouped_order.id, dp_id))

                    # Mark the order as served
                    served_orders.add(grouped_order.id)
                    remaining_orders.remove(grouped_order.id)

                    # Decrease capacity of the locker/store
                    problem.dict_capacity[dp_id] -= 1

        # Decrease the capacity for the selected point of delivery
        problem.dict_capacity[dp_id] -= 1

    return initial_solution


def plot_clusters_with_depot(problem, clustered_solution, plot_name="kmeans_clusters"):
    """
    Graficar los clusters con el depot, diferenciando lockers, tiendas y hogares, 
    y generar un gráfico separado solo para la leyenda.

    Parameters:
        problem : Problem
            The problem instance containing delivery points.
        clustered_solution : dict
            Dictionary with vehicle ID as key and list of delivery points as values.
        plot_name : str
            Name for the output plot file (default is 'kmeans_clusters').
    """
    fig, ax = plt.subplots(figsize=(10, 7))
    
    # Almacenar handles y labels para la leyenda
    handles_labels = {}
    
    # Extraer la coordenada del depot
    depot_coords = (problem.depot.loc.x, problem.depot.loc.y)
    sc4 = ax.scatter(depot_coords[0], depot_coords[1], color='black', marker='D', s=100, zorder=5)
    handles_labels['Depot'] = sc4

    # Colores para cada cluster
    num_clusters = len(clustered_solution)
    colors = cm.get_cmap("Set1", num_clusters)

    

    # Graficar cada cluster
    for vehicle_idx, route in clustered_solution.items():
        locker_coords = []
        shop_coords = []
        home_coords = []

        for order_id, dp_id in route:
            dp = problem.find_dp_by_id(dp_id)
            if dp_id.startswith('L'):
                locker_coords.append((dp.loc.x, dp.loc.y))
            elif dp_id.startswith('S'):
                shop_coords.append((dp.loc.x, dp.loc.y))
            else:  # Para ubicaciones que no sean lockers o tiendas, como hogares
                home_coords.append((dp.loc.x, dp.loc.y))

        # Convertir a arrays para graficar
        locker_coords = np.array(locker_coords)
        shop_coords = np.array(shop_coords)
        home_coords = np.array(home_coords)

        # Graficar lockers, tiendas y hogares para el cluster actual
        if len(locker_coords) > 0:
            sc1 = ax.scatter(locker_coords[:, 0], locker_coords[:, 1], color='#7C92F3', marker='p', s=75)
            handles_labels[f'Lockers (Cluster {vehicle_idx+1})'] = sc1

        if len(shop_coords) > 0:
            sc2 = ax.scatter(shop_coords[:, 0], shop_coords[:, 1], color='#006400', marker='H', s=75)
            handles_labels[f'Shops (Cluster {vehicle_idx+1})'] = sc2

        if len(home_coords) > 0:
            sc3 = ax.scatter(home_coords[:, 0], home_coords[:, 1], color=colors(vehicle_idx), marker='o', s=20)
            handles_labels[f'Home Delivery Points (Cluster {vehicle_idx+1})'] = sc3

    ax.set_xlabel('X Coordinate')
    ax.set_ylabel('Y Coordinate')
    ax.set_title('Clustered Delivery Points with Depot')
    ax.grid(True)

    # Guardar el gráfico principal
    fig.savefig(f"{plot_name}.png", format="png")
    plt.show()

    # Crear una figura separada solo para la leyenda
    fig_legend = plt.figure(figsize=(3, 2))
    ax_legend = fig_legend.add_subplot(111)
    ax_legend.axis("off")

    # Agregar la leyenda a la nueva figura
    ax_legend.legend(handles_labels.values(), handles_labels.keys(), loc="center", frameon=True)

    # Guardar la leyenda como imagen separada
    fig_legend.savefig(f"{plot_name}_legend.png", dpi=300, bbox_inches="tight")
    plt.show()




def plot_vehicle_routes(problem, final_routes):
    """
    Graficar las rutas de los vehículos diferenciando lockers, tiendas y hogares, 
    y generar un gráfico separado solo para la leyenda.

    Parameters:
        problem : Problem
            The problem instance containing delivery points and coordinates.
        final_routes : list
            List of routes assigned to each vehicle.
    """
    fig, ax = plt.subplots(figsize=(12, 10))
    
    # Coordenadas del depot
    depot_coords = problem.dict_xy['DEPOT']
    
    # Colores para los vehículos
    num_vehicles = len(final_routes)
    colors = cm.get_cmap('tab10', num_vehicles)
    
    # Listas para almacenar coordenadas por tipo de ubicación
    locker_coords = []
    shop_coords = []
    home_coords = []
    
    # Diccionario para almacenar los elementos de la leyenda sin duplicados
    handles_labels = {}
    
    

    # Procesar las rutas de cada vehículo
    for vehicle_idx, route in enumerate(final_routes):
        x_coords = []
        y_coords = []
        
        # Agregar coordenadas del depot al inicio
        x_coords.append(depot_coords[0])
        y_coords.append(depot_coords[1])
        
        for order_id, dp_id in route:
            # Verificar si el punto de entrega está en el diccionario de coordenadas
            if dp_id in problem.dict_xy:
                dp_coords = problem.dict_xy[dp_id]
                x_coords.append(dp_coords[0])
                y_coords.append(dp_coords[1])
                
                # Clasificar el tipo de ubicación según el prefijo del ID
                if dp_id.startswith('L'):
                    locker_coords.append(dp_coords)
                elif dp_id.startswith('S'):
                    shop_coords.append(dp_coords)
                else:
                    home_coords.append(dp_coords)

        # Volver al depot al final
        x_coords.append(depot_coords[0])
        y_coords.append(depot_coords[1])
        
        # Graficar la ruta del vehículo
        vehicle_plot, = ax.plot(x_coords, y_coords, label=f'Vehicle {vehicle_idx + 1}', color=colors(vehicle_idx))
        handles_labels[f'Vehicle {vehicle_idx + 1}'] = vehicle_plot

    # Graficar los puntos según el tipo de entrega
    if locker_coords:
        locker_coords = np.array(locker_coords)
        scatter_lockers = ax.scatter(locker_coords[:, 0], locker_coords[:, 1], marker='s', color='blue')
        handles_labels['Lockers'] = scatter_lockers

    if shop_coords:
        shop_coords = np.array(shop_coords)
        scatter_shops = ax.scatter(shop_coords[:, 0], shop_coords[:, 1], marker='^', color='green')
        handles_labels['Shops'] = scatter_shops

    if home_coords:
        home_coords = np.array(home_coords)
        scatter_homes = ax.scatter(home_coords[:, 0], home_coords[:, 1], marker='o', color='purple')
        handles_labels['Homes'] = scatter_homes

    # Graficar el depot
    depot_plot = ax.scatter(depot_coords[0], depot_coords[1], color='black', marker='D', s=100, zorder=5)
    handles_labels['Depot'] = depot_plot

    # Configuraciones del gráfico
    ax.set_title('Rutas de los vehículos con diferentes tipos de entrega')
    ax.set_xlabel('Coordenada X')
    ax.set_ylabel('Coordenada Y')
    ax.grid(True)

    # Guardar el gráfico con nombre único
    save_path = "./Imagenes/rutas.png"
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    unique_filename = f"{save_path}_{timestamp}.png"
    fig.savefig(unique_filename)
    plt.show()

    # Crear una figura separada solo para la leyenda
    fig_legend = plt.figure(figsize=(3, 2))
    ax_legend = fig_legend.add_subplot(111)
    ax_legend.axis("off")

    # Agregar la leyenda a la nueva figura
    ax_legend.legend(handles_labels.values(), handles_labels.keys(), loc="center", frameon=True)

    # Guardar la leyenda como imagen separada
    legend_filename = f"./Imagenes/legend_{timestamp}.png"
    fig_legend.savefig(legend_filename, dpi=300, bbox_inches="tight")
    plt.show()


def calculate_delivery_times(final_routes, problem, depot_start_time=None):
    """
    Calculate the exact delivery times for each order, following the logic in assign_deliveries_to_vehicles.

    Parameters:
    - final_routes: list of lists, where each sublist represents a route with tuples (order_id, location_id).
    - problem: object containing real distances, time constraints, and delivery times.
    - depot_start_time: float, optional, if not provided, uses problem.dict_twe['DEPOT'].

    Returns:
    - delivery_times: list of lists with tuples (vehicle_id, order_id, location_id, arrival_time, end_time).
    """
    delivery_times = []
    delivery_time_mapping = {
        'H': problem.dict_delivery_time['HOME'],
        'L': problem.dict_delivery_time['LOCKER'],
        'S': problem.dict_delivery_time['SHOP']
    }

    for vehicle_id, route in enumerate(final_routes):
        route_times = []
        current_location = 'DEPOT'
        current_time = depot_start_time if depot_start_time else problem.dict_twe['DEPOT']

        for order_id, dp_id in route:
            # Compute travel time using real distances
            distance = problem.dict_distance.get((current_location, dp_id), 10)  # Default to 10 if missing
            travel_time = (distance / problem.truck_speed) * 60  # Convert to minutes

            # Compute arrival time
            arrival_time = current_time + travel_time

            # Adjust for time window constraints
            waiting_time = 0
            if isinstance(problem.dict_twe[dp_id], list):  # Multiple time windows (Stores 'S')
                valid_window = False
                for twe, twl in zip(problem.dict_twe[dp_id], problem.dict_twl[dp_id]):
                    if twe <= arrival_time <= twl:
                        valid_window = True
                        break  # Found a valid window, no need to continue checking
                    elif arrival_time < twe:
                        waiting_time = twe - arrival_time  # Wait until the window opens
                        valid_window = True
                        break

                if not valid_window:
                    continue  # Skip delivery if it doesn't fit any window
            else:  # Single time window (Homes 'H' or Lockers 'L')
                twe, twl = problem.dict_twe[dp_id], problem.dict_twl[dp_id]
                if arrival_time < twe:
                    waiting_time = twe - arrival_time  # Wait until the window opens
                elif arrival_time > twl:
                    continue  # Skip delivery if it is too late

            # Get delivery time based on location type
            delivery_time = delivery_time_mapping.get(dp_id[0], problem.dict_delivery_time['DEFAULT'])

            # Compute end time
            end_time = arrival_time + waiting_time + delivery_time

            # Store the result
            route_times.append((vehicle_id, order_id, dp_id, arrival_time, end_time))

            # Update current time and location
            current_time = end_time
            current_location = dp_id

        delivery_times.append(route_times)

    return delivery_times



def plot_gantt_chart(delivery_times):
    """
    Plot an improved Gantt chart with better spacing and visualization.

    Parameters:
    - delivery_times: list of lists, where each sublist represents a vehicle's route with tuples 
      (vehicle_id, order_id, location_id, arrival_time, end_time).
    """
    fig, ax = plt.subplots(figsize=(14, 8))

    # Get unique vehicle count for coloring
    num_vehicles = len(delivery_times)
    colors = plt.cm.get_cmap("tab10", num_vehicles)

    # Add spacing factor for better separation
    y_spacing = 0.6  # Adjust this to control vertical separation

    for vehicle_id, route in enumerate(delivery_times):
        y_position = vehicle_id * y_spacing  # Increase space between vehicles
        for _, _, loc_id, start_time, end_time in route:
            ax.barh(y_position, end_time - start_time, left=start_time, 
                    color=colors(vehicle_id), edgecolor="black", height=0.5, label=f"{loc_id}" if vehicle_id == 0 else "")
            
            # Add text labels for better visibility
            # ax.text(start_time + (end_time - start_time) / 2, y_position, loc_id, 
            #         va='center', ha='center', fontsize=8, color='white', fontweight='bold')

    ax.set_xlabel("Time (minutes from depot start)")
    ax.set_ylabel("Vehicle Routes")
    ax.set_title("Delivery Schedule")
    ax.set_yticks([i * y_spacing for i in range(num_vehicles)])
    ax.set_yticklabels([f"Vehicle {i}" for i in range(num_vehicles)])
    plt.grid(axis="x", linestyle="--", alpha=0.6)
    plt.show()





